import os
import fnmatch
import shutil
import openpyxl as op
from pathlib import Path
from staticdata import *

SCRIPT_DIR = str(Path().resolve())
temp_file_path = "{}/{}".format(SCRIPT_DIR, "temp.xlsx")

result_data = {
    SpendType.Income: dict(),
    SpendType.Outcome: dict()
}


def find_files_by_pattern(directory, pattern):
    matched_files = []
    try:
        for filename in os.listdir(directory):
            if fnmatch.fnmatch(filename, pattern) and os.path.isfile(os.path.join(directory, filename)):
                matched_files.append(f'{directory}/{filename}')
    except Exception:
        pass
    return matched_files


def analyze_sheet(sheet, spend_type: SpendType):
    offset = 0 if spend_type == SpendType.Income else 1
    max_row = 50 if spend_type == SpendType.Income else 100
    result = result_data[spend_type]

    month_number = int(sheet.title.replace("–ê", ""))
    result[month_number] = dict()
    descr_index = offset * 8 + 3
    category_index = offset * 8 + 5
    amount_index = offset * 8 + 8
    for i in range(26, max_row):
        category = sheet.cell(i, category_index).value
        if category is None or category not in all_categories[spend_type]:
            continue
        description = sheet.cell(i, descr_index).value
        amount = sheet.cell(i, amount_index).value
        if amount in [0, None]:
            continue
        if category not in result[month_number]:
            result[month_number][category] = list()

        result[month_number][category].append(Spend(month_number, description, category, amount))


def calculate_summ(spend_type: SpendType):
    result = result_data[spend_type]
    result_summ = dict()
    for month in result.keys():
        for each_category in result[month]:
            if each_category not in result_summ:
                result_summ[each_category] = 0

            for each_record in result[month][each_category]:
                result_summ[each_category] = result_summ[each_category] + each_record.amount

    return result_summ


def main(person: Person):
    files = find_files_by_pattern(SOURCE_DATA_DIR, FILENAME_PATTERN)
    if len(files) != 1:
        raise RuntimeError("Found 0 or more than one file")

    shutil.copy2(files[0], temp_file_path)
    wb = op.load_workbook(filename=temp_file_path, data_only=True)
    for sheet_name in sheets[person]:
        try:
            sheet = wb[sheet_name]
            analyze_sheet(sheet, SpendType.Income)
            analyze_sheet(sheet, SpendType.Outcome)
        except Exception as ex:
            print(ex)

    print(f"\n{100 * '='}\n")

    result_summ = calculate_summ(SpendType.Income)
    for each_cat in all_categories[SpendType.Income]:
        if each_cat not in result_summ:
            continue
        print(f"{each_cat}: {result_summ[each_cat]}")
    print(f"\n{100*'='}\n")
    result_summ = calculate_summ(SpendType.Outcome)
    for each_cat in all_categories[SpendType.Outcome]:
        if each_cat not in result_summ:
            continue
        print(f"{each_cat}: {result_summ[each_cat]}")


main(Person.One)
